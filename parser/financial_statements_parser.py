# _*_ coding: UTF-8 _*_

import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import urllib.parse


def write(table, stock):
    workbook.create_sheet(stock)
    sheet = workbook.get_sheet_by_name(stock)

    current_row = 1
    for k, v in enumerate(table.find('thead').find_all('th')):
        sheet.cell(row=current_row, column=k + 1).value = v.text.strip()

    current_row = current_row + 1
    for i, j in enumerate(table.find('tbody').find_all('tr')):
        sheet.cell(row=current_row + i, column=1).value = j.find('th').text.strip()

        for k, v in enumerate(j.find_all('td')):
            sheet.cell(row=current_row + i, column=k + 2).value = v.text.strip()

    dims = {}
    offset = 1.2
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        sheet.column_dimensions[get_column_letter(col)].width = value * offset

    workbook.save('../resources/stock_10years.xlsx'.format(stock))


def parse(stock):
    link = url(stock)
    res = requests.get(link)
    res.encoding = 'euc-kr'
    html = res.text
    soup = BeautifulSoup(html, 'html.parser')

    print("========= START PARSING {} =========".format(stock))
    print(link)

    financial_year = soup.find("div", id="indexTable2")
    if financial_year:
        write(financial_year.find('table'), stock)


def url(stock):
    encode_stock = encode(stock)
    return "http://search.itooza.com/search.htm?seName={}" \
           "&jl=&search_ck=&sm=&sd=&ed=&ds_de=&page=&cpv=#indexTable2".format(encode_stock)


def encode(stock):
    return urllib.parse.quote_plus(stock, encoding='euc-kr')


if __name__ == '__main__':
    f = open('../resources/stock_list.txt', encoding='utf-8')
    workbook = Workbook()

    for stock in f:
        parse(stock.strip())
        time.sleep(3)

    # delete default sheet
    workbook.remove(workbook['Sheet'])
